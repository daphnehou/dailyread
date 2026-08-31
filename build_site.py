#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build a self-contained, mobile-friendly HTML page from io_psych_digest.xlsx.

Usage:
    python3 build_site.py            # rebuild index.html
    python3 build_site.py --push     # rebuild, then git commit + push if configured

The output (index.html) is fully self-contained: no CDN, no external fetch.
That means it works on GitHub Pages, from a local file://, and offline.
"""

import argparse
import datetime as dt
import html
import json
import os
import subprocess
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(os.path.dirname(HERE), "io_psych_digest.xlsx")
OUT_PATH = os.path.join(HERE, "index.html")

FIELDS = [
    "date", "topic", "authors", "year", "title", "journal", "cite",
    "doi", "summary", "relevance", "contribution", "uses", "cluster",
]

# The per-run "Topic Area" labels are deliberately specific (268 of the first
# 330 were unique), which makes them useless as a filter. So we roll them up
# into a small set of broad themes for browsing, and keep the original label
# as the chip on each card.
#
# Matching runs against topic + title only. The daily cluster note is
# excluded on purpose: it describes the whole run, so including it tagged
# every article in a day with every theme mentioned in that day's blurb.
#
# Articles can carry several themes — an AI-scored interview paper should be
# findable under both "AI & Future of Work" and "Assessment & Selection".
THEME_RULES = [
    ("AI & Future of Work", [
        "artificial intelligence", "ai-", "ai ", "ai,", "generative", "genai",
        "llm", "large language", "automation", "automat", "robot", "human-ai",
        "human–ai", "digital transformation", "future of work", "augmentation",
        "chatbot", "industry 4.0", "stara", "machine learning"]),
    ("Algorithmic Management & Surveillance", [
        "algorithmic", "algorithm", "surveillance", "electronic monitoring",
        "monitoring", "contested terrain", "worker control"]),
    ("Assessment & Selection", [
        "assessment", "selection", "interview", " avi", "avis", "game-based",
        "gamified", "gamification", "game-related", "gba", "hiring", "recruit",
        "applicant", "test taker", "personnel select"]),
    ("Careers & Vocational Behavior", [
        "career", "vocational", "employability", "calling", "decent work",
        "job search", "job seeking", "job mobility", "socialization", "newcomer",
        "person-environment fit", "interest fit", "interests", "mentoring",
        "work volition", "meaning of work", "meaningful work", "work orientation",
        "protean", "boundaryless"]),
    ("Turnover & Retention", [
        "turnover", "retention", "embeddedness", "quiet quitting", "withdrawal",
        "quitting"]),
    ("Well-Being & Occupational Health", [
        "well-being", "wellbeing", "burnout", "recovery", "detachment", "stress",
        "mental health", "exhaustion", "loneliness", "incivility", "mistreatment",
        "emotional labor", "occupational health", "worker health", "fatigue",
        "strain", "mindfulness", "safety", "sleep", "insomnia", "actigraph",
        "circadian", "shift work"]),
    ("Technostress & Digital Demands", [
        "technostress", "telepressure", "e-mail", "email", "videoconference",
        "virtual meeting", "zoom", "digital demands", "cyber incivility",
        "cyberincivility", "supplemental work", "disconnection"]),
    ("Remote & Hybrid Work", [
        "remote", "hybrid", "telework", "flexible work", "work-life",
        "work-nonwork", "boundary management", "four-day", "4-day", "virtual team"]),
    ("Training & Development", [
        "training", "learning", "upskilling", "lifelong", "transfer of training",
        "coaching", "instruction", "skill develop", "workforce development",
        "competence", "deliberate practice"]),
    ("Aging & Generations", [
        "aging", "ageism", "older worker", "generational", "multigenerational",
        "retirement", "age diver", "age-diff", "age-inclusive", "lifespan",
        "bridge employment"]),
    ("Diversity, Equity & Inclusion", [
        "diversity", "equity", "inclusion", "discrimination", "gender",
        "disability", "backlash", "inequality", "fairness", "adverse impact"]),
    ("Leadership", ["leadership", "leader"]),
    ("Teams & Collaboration", [
        "team", "collaboration", "work group", "psychological safety", "friendship"]),
    ("Employment Relationships & Gig Work", [
        "gig", "platform", "employment relationship", "psychological contract",
        "precarious", "multiple jobholding", "side-hustle", "overqualification",
        "underemployment", "non-standard", "skills mismatch", "talent management",
        "pay transparency", "compensation", "job insecurity"]),
    ("Job Design, Crafting & Motivation", [
        "job crafting", "work design", "job demands", "motivation", "engagement",
        "voice", "silence", "green behavior", "sustainab", "proactive",
        "performance management", "performance appraisal", "performance evaluation",
        "feedback", "change management", "organizational change", "passion",
        "job satisfaction"]),
    ("Big Data & Methods", [
        "methods", "measurement", "scale development", "text analysis",
        "text mining", "big data", "digital traces", "social media",
        "psychometric", "research method", "simulator", "explainability"]),
]


def themes_for(rec):
    blob = (" %s | %s " % (rec["topic"], rec["title"])).lower()
    found = [name for name, kws in THEME_RULES if any(k in blob for k in kws)]
    return found or ["Other"]


# --------------------------------------------------------------------------
# Data
# --------------------------------------------------------------------------

def norm_date(v):
    """Normalise a cell value to YYYY-MM-DD, or '' if unparseable."""
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def clean(v):
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").strip()


def read_articles(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        vals = list(r) + [None] * (len(FIELDS) - len(r))
        rec = {k: clean(v) for k, v in zip(FIELDS, vals[:len(FIELDS)])}
        rec["date"] = norm_date(vals[0])
        if not rec["title"]:
            continue
        rec["themes"] = themes_for(rec)
        rows.append(rec)
    wb.close()
    # Newest first; stable within a day so spreadsheet order is preserved.
    rows.sort(key=lambda x: x["date"], reverse=True)
    return rows


# --------------------------------------------------------------------------
# Template
# --------------------------------------------------------------------------

TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="color-scheme" content="light dark">
<meta name="theme-color" content="#1f4634" media="(prefers-color-scheme: light)">
<meta name="theme-color" content="#16291f" media="(prefers-color-scheme: dark)">
<meta name="description" content="A running digest of peer-reviewed I-O psychology research.">
<title>I-O Psych Digest</title>
<link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>&#128218;</text></svg>">
<style>
:root{
  --bg:#fdf5f7; --card:#ffffff; --ink:#17241d; --muted:#5a6f63; --faint:#6f8177;
  --line:#eddde3; --header-bg:#1f4634; --header-ink:#fbe9ef;
  --accent:#1f4634; --accent-ink:#ffffff; --chip:#fae3ea; --chip-ink:#2c5844;
  --shadow:0 1px 2px rgba(31,70,52,.06),0 1px 3px rgba(31,70,52,.06);
  --hl:#f9c8d9; --hl-ink:#4a1f2f;
}
@media (prefers-color-scheme:dark){
  :root{
    --bg:#0d1a14; --card:#15241b; --ink:#eadfe3; --muted:#9db0a5; --faint:#6d8175;
    --line:#213328; --header-bg:#16291f; --header-ink:#f5d9e2;
    --accent:#f2b6c9; --accent-ink:#14261c; --chip:#20342a; --chip-ink:#f0bfd0;
    --shadow:none; --hl:#7a3b52; --hl-ink:#ffdfe9;
  }
}
*{box-sizing:border-box;-webkit-tap-highlight-color:transparent}
html{-webkit-text-size-adjust:100%}
body{
  margin:0;background:var(--bg);color:var(--ink);
  font:16px/1.55 -apple-system,BlinkMacSystemFont,"SF Pro Text","Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;
  padding-bottom:env(safe-area-inset-bottom);
  overflow-wrap:break-word;
}
.wrap{max-width:820px;margin:0 auto;padding:0 14px}

/* ---------- header ---------- */
header{background:var(--header-bg);color:var(--header-ink);padding:22px 0 18px;padding-top:calc(22px + env(safe-area-inset-top))}
header h1{margin:0;font-size:1.45rem;letter-spacing:-.02em;font-weight:650}
header p{margin:5px 0 0;font-size:.86rem;opacity:.82}
.stats{margin-top:12px;display:flex;gap:16px;flex-wrap:wrap;font-size:.78rem;opacity:.9}
.stats b{font-size:1.02rem;font-weight:650;display:block;line-height:1.2}

/* ---------- controls ---------- */
.controls{position:sticky;top:0;z-index:20;background:var(--bg);border-bottom:1px solid var(--line);padding:9px 0 8px}
.controls .wrap{display:flex;gap:7px;flex-wrap:wrap}
input[type=search],select{
  font:inherit;font-size:16px;color:var(--ink);background:var(--card);
  border:1px solid var(--line);border-radius:10px;padding:11px 12px;min-height:44px;
  -webkit-appearance:none;appearance:none;width:100%
}
input[type=search]{flex:1 1 100%}
select{background-image:linear-gradient(45deg,transparent 50%,var(--muted) 50%),linear-gradient(135deg,var(--muted) 50%,transparent 50%);background-position:calc(100% - 17px) 20px,calc(100% - 12px) 20px;background-size:5px 5px,5px 5px;background-repeat:no-repeat;padding-right:34px;text-overflow:ellipsis}
/* Journal names run to 78 chars, so that one gets its own row. */
#jr{flex:1 1 100%}
#topic{flex:1 1 calc(58% - 4px)}
#yr{flex:1 1 calc(42% - 4px)}
input:focus,select:focus,button:focus-visible{outline:2px solid var(--accent);outline-offset:1px}
.count{font-size:.8rem;color:var(--muted);padding:2px 0 0;flex:1 1 100%}
.count button{background:none;border:none;color:var(--accent);font:inherit;font-size:.8rem;padding:0 0 0 8px;cursor:pointer;text-decoration:underline}

/* ---------- feed ---------- */
main{padding:6px 0 48px}
.daygroup{margin-top:22px}
.dayhead{
  position:sticky;top:var(--ctrl-h,150px);z-index:10;display:flex;align-items:baseline;gap:9px;
  padding:7px 0 7px;margin-bottom:9px;background:var(--bg);border-bottom:1px solid var(--line)
}
.dayhead h2{margin:0;font-size:.95rem;font-weight:650;letter-spacing:-.01em}
.dayhead span{font-size:.75rem;color:var(--faint)}
.cluster{font-size:.79rem;color:var(--muted);margin:-2px 0 12px;font-style:italic;line-height:1.45}

article{
  background:var(--card);border:1px solid var(--line);border-radius:13px;
  margin-bottom:9px;box-shadow:var(--shadow);overflow:hidden
}
.head{padding:13px 14px;cursor:pointer;display:block;user-select:none}
.chip{
  display:inline-block;background:var(--chip);color:var(--chip-ink);font-size:.685rem;
  font-weight:600;letter-spacing:.02em;padding:3px 8px;border-radius:99px;margin-bottom:7px
}
.t{font-size:.985rem;font-weight:600;line-height:1.34;letter-spacing:-.01em;margin:0 0 5px}
.m{font-size:.795rem;color:var(--muted);line-height:1.45}
.m .j{font-style:italic}
.more{font-size:.735rem;color:var(--accent);margin-top:7px;font-weight:600}
.more::after{content:" ▾"}
article.open .more::after{content:" ▴"}
.body{display:none;padding:0 14px 14px;border-top:1px solid var(--line);margin-top:2px;padding-top:12px}
article.open .body{display:block}
.body h3{margin:0 0 3px;font-size:.7rem;text-transform:uppercase;letter-spacing:.07em;color:var(--faint);font-weight:700}
.body p{margin:0 0 13px;font-size:.875rem;line-height:1.6}
.body p:last-of-type{margin-bottom:11px}
.doi{
  display:inline-block;background:var(--accent);color:var(--accent-ink);text-decoration:none;
  font-size:.79rem;font-weight:600;padding:10px 15px;border-radius:9px;min-height:40px;line-height:20px
}
mark{background:var(--hl);color:var(--hl-ink);border-radius:2px;padding:0 1px}
.empty{text-align:center;color:var(--muted);padding:52px 16px;font-size:.9rem}
footer{border-top:1px solid var(--line);padding:20px 0 34px;font-size:.75rem;color:var(--faint);text-align:center}
#sentinel{height:1px}
@media (prefers-reduced-motion:no-preference){article{transition:border-color .15s}}
</style>
</head>
<body>

<header>
  <div class="wrap">
    <h1>I-O Psych Digest</h1>
    <p>Peer-reviewed research on work, careers, technology &amp; well-being</p>
    <div class="stats">
      <div><b id="s-art">—</b>articles</div>
      <div><b id="s-day">—</b>days</div>
      <div><b id="s-top">—</b>themes</div>
      <div><b id="s-upd">—</b>last entry</div>
    </div>
  </div>
</header>

<div class="controls" id="controls">
  <div class="wrap">
    <input type="search" id="q" placeholder="Search titles, authors, journals, summaries…" autocomplete="off" autocorrect="off" spellcheck="false" enterkeyhint="search">
    <select id="jr"><option value="">All journals</option></select>
    <select id="topic"><option value="">All themes</option></select>
    <select id="yr"><option value="">All years</option></select>
    <div class="count"><span id="count"></span><button type="button" id="reset" hidden>Clear filters</button></div>
  </div>
</div>

<main class="wrap" id="feed"></main>
<div id="sentinel"></div>
<footer class="wrap">Built <span id="built"></span> · source: io_psych_digest.xlsx</footer>

<script type="application/json" id="data">__DATA__</script>
<script>
(function(){
"use strict";
var DATA  = JSON.parse(document.getElementById("data").textContent);
var BUILT = "__BUILT__";
document.getElementById("built").textContent = BUILT;

var $ = function(id){ return document.getElementById(id); };
var feed = $("feed"), qEl = $("q"), topicEl = $("topic"), yrEl = $("yr"), jrEl = $("jr");
var countEl = $("count"), resetEl = $("reset");

/* ---------- searchable blob, built once ---------- */
DATA.forEach(function(a, i){
  a._i = i;
  a._s = [a.title,a.authors,a.journal,a.topic,a.summary,a.relevance,
          a.contribution,a.uses,a.year,a.cite,(a.themes||[]).join(" ")]
          .join(" ").toLowerCase();
});

/* ---------- header stats ---------- */
var days = {}, topics = {}, years = {}, journals = {};
DATA.forEach(function(a){
  if(a.date)  days[a.date]  = 1;
  if(a.year)  years[a.year] = 1;
  if(a.journal) journals[a.journal] = (journals[a.journal] || 0) + 1;
  (a.themes || []).forEach(function(t){ topics[t] = 1; });
});
var dayKeys = Object.keys(days).sort().reverse();
$("s-art").textContent = DATA.length;
$("s-day").textContent = dayKeys.length;
$("s-top").textContent = Object.keys(topics).length;
$("s-upd").textContent = dayKeys.length ? fmtShort(dayKeys[0]) : "—";

Object.keys(topics).sort().forEach(function(t){
  var o = document.createElement("option"); o.value = t; o.textContent = t; topicEl.appendChild(o);
});
Object.keys(years).sort().reverse().forEach(function(y){
  var o = document.createElement("option"); o.value = y; o.textContent = y; yrEl.appendChild(o);
});
/* Journals are ordered by how often they appear, so the outlets that actually
   carry this literature sit at the top of the picker instead of being buried
   alphabetically among 90 entries. */
Object.keys(journals).sort(function(a, b){
  return (journals[b] - journals[a]) || a.localeCompare(b);
}).forEach(function(j){
  var o = document.createElement("option");
  o.value = j; o.textContent = j + "  (" + journals[j] + ")";
  jrEl.appendChild(o);
});

/* ---------- helpers ---------- */
function esc(s){
  return String(s == null ? "" : s)
    .replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;");
}
function fmtShort(iso){
  var p = String(iso).split("-");
  if(p.length !== 3) return iso;
  var M = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
  return M[(+p[1]) - 1] + " " + (+p[2]);
}
function fmtLong(iso){
  var p = String(iso).split("-");
  if(p.length !== 3) return iso;
  var d = new Date(+p[0], +p[1] - 1, +p[2]);
  if(isNaN(d)) return iso;
  return d.toLocaleDateString(undefined,{weekday:"short",month:"long",day:"numeric",year:"numeric"});
}
/* Authors are stored "Last, F.; Last, F." — shorten for the collapsed card. */
function shortAuthors(s){
  if(!s) return "";
  var parts = s.split(";").map(function(x){ return x.trim(); }).filter(Boolean);
  if(parts.length === 0) return s;
  var first = parts[0].split(",")[0].trim();
  if(parts.length === 1) return first;
  if(parts.length === 2) return first + " & " + parts[1].split(",")[0].trim();
  return first + " et al.";
}
function rx(term){
  return new RegExp("(" + term.replace(/[.*+?^${}()|[\]\\]/g,"\\$&") + ")","gi");
}
function hl(text, term){
  var e = esc(text);
  if(!term) return e;
  return e.replace(rx(term), "<mark>$1</mark>");
}

/* ---------- rendering (batched so long lists stay snappy) ---------- */
var visible = [], drawn = 0, term = "", BATCH = 40, lastDay = null, curGroup = null;

function cardHTML(a){
  var meta = [shortAuthors(a.authors), a.year].filter(Boolean).join(" · ");
  return '<div class="head" data-i="' + a._i + '">' +
           (a.topic ? '<span class="chip">' + esc(a.topic) + "</span>" : "") +
           '<p class="t">' + hl(a.title, term) + "</p>" +
           '<div class="m">' + hl(meta, term) +
             (a.journal ? ' · <span class="j">' + hl(a.journal, term) + "</span>" : "") +
             (a.cite ? ", " + esc(a.cite) : "") +
           "</div>" +
           '<div class="more">Details</div>' +
         "</div>" +
         '<div class="body" data-loaded="0"></div>';
}

function bodyHTML(a){
  var out = "";
  if(a.summary)      out += "<h3>Summary</h3><p>"       + hl(a.summary, term)      + "</p>";
  if(a.relevance)    out += "<h3>Relevance</h3><p>"     + hl(a.relevance, term)    + "</p>";
  if(a.contribution) out += "<h3>Contribution</h3><p>"  + hl(a.contribution, term) + "</p>";
  if(a.uses)         out += "<h3>Use cases</h3><p>"     + hl(a.uses, term)         + "</p>";
  if(a.doi){
    var href = /^https?:/i.test(a.doi) ? a.doi : "https://doi.org/" + String(a.doi).replace(/^doi:\s*/i,"");
    out += '<a class="doi" href="' + esc(href) + '" target="_blank" rel="noopener noreferrer">Open article ↗</a>';
  }
  return out || "<p>No further detail recorded.</p>";
}

function drawBatch(){
  var slice = visible.slice(drawn, drawn + BATCH);
  if(!slice.length) return;
  var frag = document.createDocumentFragment();

  slice.forEach(function(a){
    /* curGroup persists across batches: a day's articles can straddle a
       batch boundary, in which case we keep filling the existing section. */
    if(a.date !== lastDay || !curGroup){
      lastDay = a.date;
      curGroup = document.createElement("section");
      curGroup.className = "daygroup";
      var h = '<div class="dayhead"><h2>' + esc(fmtLong(a.date)) + "</h2>" +
              "<span>" + countForDay(a.date) + "</span></div>";
      if(a.cluster) h += '<p class="cluster">' + esc(a.cluster) + "</p>";
      curGroup.innerHTML = h;
      frag.appendChild(curGroup);
    }
    var art = document.createElement("article");
    art.innerHTML = cardHTML(a);
    curGroup.appendChild(art);
  });

  feed.appendChild(frag);
  drawn += slice.length;
}

function countForDay(d){
  var n = visible.filter(function(a){ return a.date === d; }).length;
  return n + (n === 1 ? " entry" : " entries");
}

function render(){
  var t  = qEl.value.trim().toLowerCase();
  var tp = topicEl.value, y = yrEl.value, j = jrEl.value;
  term = t;

  visible = DATA.filter(function(a){
    if(tp && (a.themes || []).indexOf(tp) === -1) return false;
    if(y  && String(a.year) !== y) return false;
    if(j  && a.journal !== j) return false;
    if(t  && a._s.indexOf(t) === -1) return false;
    return true;
  });

  feed.innerHTML = ""; drawn = 0; lastDay = null; curGroup = null;

  var filtered = !!(t || tp || y || j);
  countEl.textContent = filtered
    ? visible.length + " of " + DATA.length + " articles"
    : DATA.length + " articles";
  resetEl.hidden = !filtered;

  if(!visible.length){
    feed.innerHTML = '<p class="empty">Nothing matches that.<br>Try a shorter search term.</p>';
    return;
  }
  drawBatch();
  fill();
}

/* Keep drawing until the sentinel is pushed off-screen. */
function fill(){
  var guard = 0;
  while(drawn < visible.length && guard++ < 30 &&
        $("sentinel").getBoundingClientRect().top < window.innerHeight * 2){
    drawBatch();
  }
}

/* ---------- events ---------- */
feed.addEventListener("click", function(e){
  var head = e.target.closest(".head");
  if(!head) return;
  var art  = head.parentNode;
  var body = head.nextElementSibling;
  if(body.dataset.loaded === "0"){
    body.innerHTML = bodyHTML(DATA[+head.dataset.i]);
    body.dataset.loaded = "1";
  }
  art.classList.toggle("open");
});

var timer;
qEl.addEventListener("input", function(){
  clearTimeout(timer);
  timer = setTimeout(render, 140);
});
topicEl.addEventListener("change", render);
yrEl.addEventListener("change", render);
jrEl.addEventListener("change", render);
resetEl.addEventListener("click", function(){
  qEl.value = ""; topicEl.value = ""; yrEl.value = ""; jrEl.value = "";
  render(); window.scrollTo(0, 0);
});

/* The control bar wraps to a different number of rows depending on screen
   width, so the sticky day headings measure it rather than assuming a height. */
function syncStick(){
  document.documentElement.style.setProperty("--ctrl-h", $("controls").offsetHeight + "px");
}
syncStick();
window.addEventListener("resize", syncStick);
window.addEventListener("orientationchange", syncStick);

if("IntersectionObserver" in window){
  new IntersectionObserver(function(entries){
    if(entries[0].isIntersecting) fill();
  }, {rootMargin:"600px"}).observe($("sentinel"));
}else{
  window.addEventListener("scroll", fill, {passive:true});
}

render();
})();
</script>
</body>
</html>
"""


# --------------------------------------------------------------------------
# Build
# --------------------------------------------------------------------------

def build():
    if not os.path.exists(EXCEL_PATH):
        sys.exit("Spreadsheet not found: %s" % EXCEL_PATH)

    articles = read_articles(EXCEL_PATH)
    if not articles:
        sys.exit("No articles found in %s" % EXCEL_PATH)

    payload = json.dumps(articles, ensure_ascii=False, separators=(",", ":"))
    # Neutralise anything that could close the <script> tag early.
    payload = payload.replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026")

    built = dt.datetime.now().strftime("%d %b %Y, %H:%M")
    page = TEMPLATE.replace("__DATA__", payload).replace("__BUILT__", html.escape(built))

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        f.write(page)

    days = len({a["date"] for a in articles if a["date"]})
    kb = os.path.getsize(OUT_PATH) / 1024
    print("Built %s" % OUT_PATH)
    print("  %d articles across %d days  (%.0f KB)" % (len(articles), days, kb))
    return articles


# --------------------------------------------------------------------------
# Optional git push
# --------------------------------------------------------------------------

def git(*args):
    return subprocess.run(
        ["git"] + list(args), cwd=HERE,
        capture_output=True, text=True, timeout=120,
    )


def push(n_articles):
    """Commit and push. Never raises — a failed push must not fail the digest."""
    if not os.path.isdir(os.path.join(HERE, ".git")):
        print("! Not a git repo yet — skipping push. See README.md for setup.")
        return

    if git("remote").stdout.strip() == "":
        print("! No git remote configured — skipping push. See README.md.")
        return

    git("add", "-A")
    if git("diff", "--staged", "--quiet").returncode == 0:
        print("  Nothing changed — no commit needed.")
        return

    msg = "Digest update %s (%d articles)" % (dt.date.today().isoformat(), n_articles)
    c = git("commit", "-m", msg)
    if c.returncode != 0:
        print("! Commit failed:\n%s" % (c.stderr or c.stdout).strip()[:400])
        return

    # Absorb anything changed on github.com directly (web-UI edits, or the CNAME
    # file GitHub writes when you set a custom domain) so the push isn't rejected.
    git("pull", "--rebase", "origin", "main")

    p = git("push", "origin", "main")
    if p.returncode == 0:
        print("  Pushed to GitHub.")
    else:
        print("! Push failed (commit is saved locally, push it yourself):")
        print((p.stderr or p.stdout).strip()[:400])


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Build the I-O Psych Digest page.")
    ap.add_argument("--push", action="store_true", help="git commit + push after building")
    args = ap.parse_args()

    arts = build()
    if args.push:
        push(len(arts))
